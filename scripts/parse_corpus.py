# Parse the 9 corpus xlsx files -> corpus-parsed.json (clean intermediate for the TS importer).
# READ-ONLY on the spreadsheets. Run: py -X utf8 scripts/parse_corpus.py
import openpyxl, os, json, re, unicodedata

D = r'C:\Users\juang\Downloads'
CONN_FILES = [
    ('crossshelf-connections-GAMES-part1.xlsx', 'GAMES', 'game'),
    ('crossshelf-connections-GAMES-part2.xlsx', 'GAMES', 'game'),
    ('crossshelf-connections-MOVIES-part1.xlsx', 'MOVIES', 'movie'),
    ('crossshelf-connections-MOVIES-part2.xlsx', 'MOVIES', 'movie'),
    ('crossshelf-connections-BOOKS-part1.xlsx', 'BOOKS', 'book'),
    ('crossshelf-connections-BOOKS-part2.xlsx', 'BOOKS', 'book'),
    ('crossshelf-connections-TV.xlsx', 'TV', 'tv'),
    ('crossshelf-connections-ANIME.xlsx', 'ANIME', 'anime'),
]

def norm(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii')
    s = s.lower().replace('&', ' and ').replace("'", '')
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    return re.sub(r'\s+', ' ', s)

def first_line(s):
    return str(s).split('\n')[0].strip() if s else None

connections = []
for fname, ftype, anchor_cat in CONN_FILES:
    wb = openpyxl.load_workbook(os.path.join(D, fname), data_only=True)
    ws = [s for s in wb.worksheets if s.title != 'How to use'][0]
    rows = list(ws.iter_rows(values_only=True))[1:]
    anchor = cluster = blurb = None
    for r in rows:
        a, c, b, title, media, whatis, stren, cat, threads = r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]
        if a:
            anchor = first_line(a)
        if c:
            cluster = str(c).strip(); blurb = (str(b).strip() if b else '')
        if title:
            connections.append({
                'file_type': ftype,
                'anchor_title': anchor,
                'anchor_cat_type': anchor_cat,
                'cluster_label': cluster,
                'blurb': blurb,
                'rec_title': str(title).strip(),
                'rec_media': (str(media).strip() if media else None),
                'what_it_is': (str(whatis).strip() if whatis else None),
                'strength': (str(stren).strip().lower() if stren else None),
                'in_catalog_flag': (str(cat).strip() if cat else None),
                'shared_threads': [t.strip() for t in re.split(r'[·;|]', str(threads))] if threads else [],
            })
    wb.close()

# Cluster library: Canonical Clusters + Full mapping
wb = openpyxl.load_workbook(os.path.join(D, 'crossshelf-cluster-library-consolidated.xlsx'), data_only=True)
canon_rows = list(wb['Canonical Clusters'].iter_rows(values_only=True))[1:]
canonical = []
cur = None
for r in canon_rows:
    name, cblurb, spans, anchors, title, media, stren, cat = r[:8]
    if name:
        cur = {'name': str(name).strip(), 'blurb': (str(cblurb).strip() if cblurb else ''),
               'spans': (str(spans).split() if spans else []), 'slug': norm(name)}
        canonical.append(cur)

full_rows = list(wb['Full mapping'].iter_rows(values_only=True))[1:]
full_mapping = {}
for r in full_rows:
    typ, anc, orig, canon, merged = r[:5]
    if not orig:
        continue
    key = f'{str(typ).strip()}||{norm(anc)}||{norm(orig)}'
    full_mapping[key] = {'canonical': str(canon).strip(), 'canonical_slug': norm(canon), 'merged': str(merged).strip()}
wb.close()

out = {'connections': connections, 'canonical_clusters': canonical, 'full_mapping': full_mapping}
with open('corpus-parsed.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False)

# join coverage check: how many connections find their canonical via full_mapping
hit = 0
miss = 0
for cn in connections:
    key = f'{cn["file_type"]}||{norm(cn["anchor_title"])}||{norm(cn["cluster_label"])}'
    if key in full_mapping:
        hit += 1
    else:
        miss += 1
print(f'connections={len(connections)}  canonical_clusters={len(canonical)}  full_mapping_rows={len(full_mapping)}')
print(f'connection->canonical join: hit={hit} miss={miss}')
anchors = sorted({(c["anchor_title"], c["anchor_cat_type"]) for c in connections})
print(f'distinct anchors={len(anchors)}')
print('wrote corpus-parsed.json')
